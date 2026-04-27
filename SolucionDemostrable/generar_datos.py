import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Facturas"

headers = ["ID", "Cliente", "RFC", "Referencia", "Monto", "Fecha_Vencimiento", "Estado"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E79")
    cell.alignment = Alignment(horizontal="center")

facturas = [
    (1, "Empresa ABC S.A. de C.V.", "ABC010101ABC", "TRF-001", 15400.00, "2025-04-30", "pendiente"),
    (2, "Grupo XYZ Comercial",       "GXY020202GXY", "TRF-002",  8750.50, "2025-04-28", "pendiente"),
    (3, "Distribuidora Nexo",        "DNX030303DNX", "TRF-003", 32000.00, "2025-05-05", "pendiente"),
    (4, "Soluciones Tech MX",        "STM040404STM", "TRF-004",  5200.00, "2025-04-25", "pendiente"),
    (5, "Comercial del Norte",       "CDN050505CDN", "TRF-005", 11800.75, "2025-05-10", "pendiente"),
]

for factura in facturas:
    ws.append(factura)

for col in ws.columns:
    ws.column_dimensions[col[0].column_letter].width = 22

wb.save("facturas.xlsx")
print("facturas.xlsx generado correctamente")