import openpyxl

EXCEL_PATH = "facturas.xlsx"
TOLERANCIA = 0.01  # diferencia mínima aceptable en monto

def cargar_facturas() -> list:
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        
        facturas = []
        headers = [cell.value for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            factura = dict(zip(headers, row))
            facturas.append(factura)
        
        return facturas

    except Exception as e:
        raise Exception(f"No se pudo leer el Excel: {str(e)}")


def validar_pago(datos_extraidos: dict) -> dict:
    facturas = cargar_facturas()

    referencia = datos_extraidos.get("referencia", "").strip().upper()
    monto_pagado = float(datos_extraidos.get("monto", 0))
    pagador = datos_extraidos.get("pagador", "Desconocido")

    # Buscar factura por referencia
    factura_encontrada = None
    for factura in facturas:
        ref_factura = str(factura.get("Referencia", "")).strip().upper()
        if ref_factura == referencia:
            factura_encontrada = factura
            break

    # Caso 1: referencia no encontrada
    if not factura_encontrada:
        return {
            "resultado": "no_encontrada",
            "accion": "escalado",
            "mensaje": f"No se encontró factura con referencia '{referencia}'",
            "datos_pago": datos_extraidos,
            "factura": None
        }

    monto_esperado = float(factura_encontrada.get("Monto", 0))
    diferencia = abs(monto_pagado - monto_esperado)

    # Caso 2: coincidencia exacta
    if diferencia <= TOLERANCIA:
        actualizar_factura(referencia)
        return {
            "resultado": "coincidencia_exacta",
            "accion": "actualizado",
            "mensaje": f"Pago correcto. Factura {referencia} marcada como pagada.",
            "datos_pago": datos_extraidos,
            "factura": factura_encontrada
        }

    # Caso 3: pago parcial
    if monto_pagado < monto_esperado:
        return {
            "resultado": "pago_parcial",
            "accion": "escalado",
            "mensaje": f"Pago incompleto. Se esperaban ${monto_esperado:.2f}, se recibieron ${monto_pagado:.2f}. Diferencia: ${diferencia:.2f}",
            "datos_pago": datos_extraidos,
            "factura": factura_encontrada
        }

    # Caso 4: pago de más
    return {
        "resultado": "pago_excedente",
        "accion": "escalado",
        "mensaje": f"Pago excedente. Se esperaban ${monto_esperado:.2f}, se recibieron ${monto_pagado:.2f}. Diferencia: ${diferencia:.2f}",
        "datos_pago": datos_extraidos,
        "factura": factura_encontrada
    }


def actualizar_factura(referencia: str):
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        ref_cell = row[3]  # columna Referencia
        estado_cell = row[6]  # columna Estado
        if str(ref_cell.value).strip().upper() == referencia:
            estado_cell.value = "pagada"
            break

    wb.save(EXCEL_PATH)